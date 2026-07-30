"""Published trade-ledger source, validation, and export helpers.

The module deliberately stays independent from Dash so source and workbook
contracts can be tested without importing the application callback registry.
"""

from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal
import hashlib
import io
import json
import math
from typing import Any, Iterable, Mapping, Sequence

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from sqlalchemy import bindparam, text

from db_fallback import DB_SCHEMA, fq_table


VALUATION_CURRENT_TABLE = fq_table(DB_SCHEMA, "trades_options_valuation_current")
VALUATION_CURRENT_TABLE_NAME = "trades_options_valuation_current"

ORIGINAL_MODEL_GREEK_COLUMNS = [
    "delta_s1",
    "delta_s2",
    "gamma_s1",
    "gamma_s2",
    "gamma_s1s2",
    "vega_sigma1",
    "vega_sigma2",
    "corr_sensitivity",
    "theta",
    "rho",
]

ORIGINAL_POSITION_GREEK_COLUMNS = [
    "qty_delta_asset_a",
    "qty_delta_asset_b",
    "qty_gamma_asset_a",
    "qty_gamma_asset_b",
    "qty_delta_s1",
    "qty_delta_s2",
    "qty_gamma_s1",
    "qty_gamma_s2",
    "qty_gamma_s1s2",
    "qty_vega_sigma1",
    "qty_vega_sigma2",
    "qty_corr_sensitivity",
    "qty_theta",
    "qty_rho",
]

MONETARY_TOTAL_COLUMNS = {
    "qty_value",
    "qty_intrinsic_value",
    "qty_time_value",
    "qty_premium",
    "qty_pnl",
}

ORIGINAL_GREEK_SOURCE_COLUMNS = [
    '"delta_S1"',
    '"delta_S2"',
    '"gamma_S1"',
    '"gamma_S2"',
    '"gamma_S1S2"',
    "vega_sigma1",
    "vega_sigma2",
    "corr_sensitivity",
    "theta",
    "rho",
    "qty_delta_asset_a",
    "qty_delta_asset_b",
    "qty_gamma_asset_a",
    "qty_gamma_asset_b",
    '"qty_delta_S1"',
    '"qty_delta_S2"',
    '"qty_gamma_S1"',
    '"qty_gamma_S2"',
    '"qty_gamma_S1S2"',
    "qty_vega_sigma1",
    "qty_vega_sigma2",
    "qty_corr_sensitivity",
    "qty_theta",
    "qty_rho",
]

SOURCE_COLUMNS = [
    "cob_date",
    "trade_date",
    "entity",
    "type_trade",
    "book",
    "strategy",
    "substrategy",
    "type_option",
    "model",
    "put_call",
    "buy_sell",
    "currency",
    "premium",
    "expiration_date",
    "quantity",
    "unit_quantity",
    "strike",
    "asset_a",
    "asset_a_multiplier",
    "asset_a_premium",
    "maturity_date_type_a",
    "maturity_date_a",
    "asset_sign_a",
    "asset_b",
    "asset_b_multiplier",
    "asset_b_premium",
    "maturity_date_type_b",
    "maturity_date_b",
    "asset_sign_b",
    "asset_c",
    "asset_c_multiplier",
    "asset_c_premium",
    "maturity_date_type_c",
    "maturity_date_c",
    "asset_sign_c",
    "price_a",
    "price_b",
    "price_c",
    "vol_a",
    "vol_b",
    "vol_c",
    "correlation",
    "adjusted_price_a",
    "adjusted_price_b",
    "adjusted_price_c",
    "adjusted_strike",
    "time_to_expiry",
    "adjusted_vol_a",
    "adjusted_vol_b",
    "adjusted_vol_c",
    "price",
    '"delta_S1" AS delta_s1',
    '"delta_S2" AS delta_s2',
    '"gamma_S1" AS gamma_s1',
    '"gamma_S2" AS gamma_s2',
    '"gamma_S1S2" AS gamma_s1s2',
    "vega_sigma1",
    "vega_sigma2",
    "corr_sensitivity",
    "theta",
    "rho",
    "intrinsic_value",
    "time_value",
    "pnl",
    "qty_delta_asset_a",
    "qty_delta_asset_b",
    "qty_gamma_asset_a",
    "qty_gamma_asset_b",
    '"qty_delta_S1" AS qty_delta_s1',
    '"qty_delta_S2" AS qty_delta_s2',
    '"qty_gamma_S1" AS qty_gamma_s1',
    '"qty_gamma_S2" AS qty_gamma_s2',
    '"qty_gamma_S1S2" AS qty_gamma_s1s2',
    "qty_vega_sigma1",
    "qty_vega_sigma2",
    "qty_corr_sensitivity",
    "qty_theta",
    "qty_rho",
    "qty_value",
    "qty_intrinsic_value",
    "qty_time_value",
    "qty_premium",
    "qty_pnl",
    "contract_convention_code",
    "discount_curve_code",
    "margin_style",
    "discount_curve_cob_date",
    "discount_factor_to_expiry",
    "pricing_model_version",
    "convention_source_url",
    "forward_price_used",
    "volatility_used",
    "valuation_run_id",
    "valuation_revision",
    "valuation_methodology_version",
    "valuation_input_fingerprint",
    "valuation_created_at",
    "valuation_created_by",
    "valuation_published_at",
    "valuation_published_by",
]

OUTPUT_COLUMNS = [
    column.split(" AS ")[-1].strip('"')
    for column in SOURCE_COLUMNS
]

DATE_COLUMNS = {
    "cob_date",
    "trade_date",
    "expiration_date",
    "maturity_date_a",
    "maturity_date_b",
    "maturity_date_c",
    "discount_curve_cob_date",
}

TIMESTAMP_COLUMNS = {
    "valuation_created_at",
    "valuation_published_at",
}

TEXT_COLUMNS = {
    "entity",
    "type_trade",
    "book",
    "strategy",
    "substrategy",
    "type_option",
    "model",
    "put_call",
    "buy_sell",
    "currency",
    "unit_quantity",
    "asset_a",
    "maturity_date_type_a",
    "asset_sign_a",
    "asset_b",
    "maturity_date_type_b",
    "asset_sign_b",
    "asset_c",
    "maturity_date_type_c",
    "asset_sign_c",
    "contract_convention_code",
    "discount_curve_code",
    "margin_style",
    "pricing_model_version",
    "convention_source_url",
    "valuation_methodology_version",
    "valuation_input_fingerprint",
    "valuation_created_by",
    "valuation_published_by",
}

IDENTITY_COLUMNS = [
    "trade_date",
    "entity",
    "type_trade",
    "book",
    "strategy",
    "substrategy",
    "type_option",
    "model",
    "put_call",
    "buy_sell",
    "currency",
    "premium",
    "expiration_date",
    "quantity",
    "unit_quantity",
    "strike",
    "asset_a",
    "asset_a_multiplier",
    "asset_a_premium",
    "maturity_date_type_a",
    "maturity_date_a",
    "asset_sign_a",
    "asset_b",
    "asset_b_multiplier",
    "asset_b_premium",
    "maturity_date_type_b",
    "maturity_date_b",
    "asset_sign_b",
    "asset_c",
    "asset_c_multiplier",
    "asset_c_premium",
    "maturity_date_type_c",
    "maturity_date_c",
    "asset_sign_c",
    "contract_convention_code",
    "discount_curve_code",
]


class TradeLedgerDataError(RuntimeError):
    """Raised when published data cannot safely support the trade ledger."""


@dataclass(frozen=True)
class TradeSnapshot:
    rows: pd.DataFrame
    cob_date: str
    valuation_run_id: str
    valuation_revision: int
    valuation_methodology_version: str
    valuation_published_at: str
    valuation_published_by: str | None

    @property
    def row_count(self) -> int:
        return len(self.rows)

    def metadata(self) -> dict[str, Any]:
        return {
            "source": f"{DB_SCHEMA}.{VALUATION_CURRENT_TABLE_NAME}",
            "cob_date": self.cob_date,
            "valuation_run_id": self.valuation_run_id,
            "valuation_revision": self.valuation_revision,
            "valuation_methodology_version": self.valuation_methodology_version,
            "valuation_published_at": self.valuation_published_at,
            "valuation_published_by": self.valuation_published_by,
            "row_count": self.row_count,
            "currencies": sorted(
                self.rows["currency"].dropna().astype(str).unique().tolist()
            ),
        }

    def records(self) -> list[dict[str, Any]]:
        return [
            {
                column: _json_value(
                    value,
                    date_only=column in DATE_COLUMNS,
                )
                for column, value in row.items()
            }
            for row in self.rows.to_dict("records")
        ]


def _available_columns(engine) -> set[str]:
    query = text(
        """
        SELECT column_name
        FROM information_schema.columns
        WHERE table_schema = :schema
          AND table_name = :table_name
        """
    )
    columns = pd.read_sql(
        query,
        engine,
        params={"schema": DB_SCHEMA, "table_name": VALUATION_CURRENT_TABLE_NAME},
    )
    return set(columns["column_name"].astype(str))


def assert_source_schema(engine) -> None:
    if engine is None:
        raise TradeLedgerDataError("Database configuration is unavailable.")
    available = _available_columns(engine)
    required_source_columns = {
        column.split(" AS ")[0].strip('"')
        for column in SOURCE_COLUMNS
    }
    missing = sorted(required_source_columns - available)
    if missing:
        raise TradeLedgerDataError(
            "Published trade ledger schema is incomplete; missing original "
            f"columns: {', '.join(missing)}."
        )


def get_available_cob_dates(engine) -> list[str]:
    """Return published COBs with complete original valuation Greeks."""
    assert_source_schema(engine)
    complete_predicate = " AND ".join(
        f"{column} IS NOT NULL"
        for column in ORIGINAL_GREEK_SOURCE_COLUMNS
    )
    query = text(
        f"""
        SELECT cob_date
        FROM {VALUATION_CURRENT_TABLE}
        GROUP BY cob_date
        HAVING bool_and({complete_predicate})
           AND count(DISTINCT valuation_run_id) = 1
           AND count(DISTINCT valuation_revision) = 1
           AND count(DISTINCT valuation_methodology_version) = 1
        ORDER BY cob_date DESC
        """
    )
    frame = pd.read_sql(query, engine)
    return (
        pd.to_datetime(frame["cob_date"], errors="coerce")
        .dropna()
        .dt.strftime("%Y-%m-%d")
        .tolist()
    )


def get_substrategies(engine, cob_date: str) -> list[str]:
    assert_source_schema(engine)
    parsed_date = _parse_cob_date(cob_date)
    query = text(
        f"""
        SELECT DISTINCT substrategy
        FROM {VALUATION_CURRENT_TABLE}
        WHERE cob_date = :cob_date
          AND substrategy IS NOT NULL
        ORDER BY substrategy
        """
    )
    frame = pd.read_sql(query, engine, params={"cob_date": parsed_date})
    return frame["substrategy"].astype(str).tolist()


def load_trade_snapshot(
    engine,
    cob_date: str,
    substrategies: Sequence[str] | None = None,
) -> TradeSnapshot:
    """Load and validate one immutable published valuation snapshot."""
    assert_source_schema(engine)
    parsed_date = _parse_cob_date(cob_date)
    selected_substrategies = tuple(
        sorted({str(value) for value in (substrategies or []) if str(value).strip()})
    )
    selected_columns = ",\n            ".join(SOURCE_COLUMNS)
    sql = f"""
        SELECT
            {selected_columns}
        FROM {VALUATION_CURRENT_TABLE}
        WHERE cob_date = :cob_date
    """
    params: dict[str, Any] = {"cob_date": parsed_date}
    query = text(sql)
    if selected_substrategies:
        query = text(sql + "\n AND substrategy IN :substrategies").bindparams(
            bindparam("substrategies", expanding=True)
        )
        params["substrategies"] = selected_substrategies

    frame = pd.read_sql(query, engine, params=params)
    if frame.empty:
        raise TradeLedgerDataError(
            "No active valued trades are available for the selected COB and strategies."
        )
    return _validate_snapshot(frame, parsed_date)


def _parse_cob_date(value: str) -> date:
    parsed = pd.to_datetime(value, errors="coerce")
    if pd.isna(parsed):
        raise TradeLedgerDataError("The selected COB date is invalid.")
    return parsed.date()


def _validate_snapshot(frame: pd.DataFrame, cob_date: date) -> TradeSnapshot:
    data = frame.copy()
    missing = sorted(set(OUTPUT_COLUMNS) - set(data.columns))
    if missing:
        raise TradeLedgerDataError(
            f"Trade snapshot is missing selected columns: {', '.join(missing)}."
        )

    for column in DATE_COLUMNS | TIMESTAMP_COLUMNS:
        data[column] = pd.to_datetime(data[column], errors="coerce")
    for column in set(OUTPUT_COLUMNS) - DATE_COLUMNS - TIMESTAMP_COLUMNS - TEXT_COLUMNS:
        if column == "valuation_run_id":
            continue
        data[column] = pd.to_numeric(data[column], errors="coerce")

    invalid_currency = (
        data["currency"].isna()
        | ~data["currency"].astype("string").str.fullmatch(
            r"[A-Z]{3}",
            na=False,
        )
    )
    if invalid_currency.any():
        raise TradeLedgerDataError(
            "The selected published snapshot has invalid currency codes."
        )

    original_greeks = [
        *ORIGINAL_MODEL_GREEK_COLUMNS,
        *ORIGINAL_POSITION_GREEK_COLUMNS,
    ]
    greek_missing = data[original_greeks].isna().any(axis=1)
    greek_non_finite = ~np.isfinite(
        data[original_greeks].to_numpy(dtype=float)
    ).all(axis=1)
    if (greek_missing | greek_non_finite).any():
        raise TradeLedgerDataError(
            "The selected published snapshot has incomplete original valuation Greeks."
        )

    if data["cob_date"].isna().any() or not data["cob_date"].dt.date.eq(cob_date).all():
        raise TradeLedgerDataError("The source returned rows from an unexpected COB.")

    run_id = _single_value(data, "valuation_run_id")
    revision = _single_value(data, "valuation_revision")
    methodology = _single_value(data, "valuation_methodology_version")
    published_at = _single_value(data, "valuation_published_at")
    published_by = _single_value(data, "valuation_published_by", allow_none=True)

    data["_trade_key"] = data.apply(_trade_key, axis=1)
    duplicate_keys = data["_trade_key"].duplicated(keep=False)
    if duplicate_keys.any():
        raise TradeLedgerDataError(
            "Published rows do not have a unique booking identity; "
            "the ledger refuses to collapse or duplicate trades."
        )

    data = data.sort_values(
        ["trade_date", "substrategy", "expiration_date", "put_call", "strike"],
        ascending=[False, True, True, True, True],
        kind="stable",
    ).reset_index(drop=True)

    return TradeSnapshot(
        rows=data,
        cob_date=cob_date.isoformat(),
        valuation_run_id=str(run_id),
        valuation_revision=int(revision),
        valuation_methodology_version=str(methodology),
        valuation_published_at=pd.Timestamp(published_at).isoformat(),
        valuation_published_by=None if published_by is None else str(published_by),
    )


def _single_value(
    frame: pd.DataFrame,
    column: str,
    *,
    allow_none: bool = False,
) -> Any:
    values = frame[column].drop_duplicates()
    if len(values) != 1:
        raise TradeLedgerDataError(
            f"Published snapshot mixes multiple values for {column}."
        )
    value = values.iloc[0]
    if pd.isna(value):
        if allow_none:
            return None
        raise TradeLedgerDataError(
            f"Published snapshot is missing required metadata: {column}."
        )
    return value


def _identity_value(value: Any) -> Any:
    if value is None or pd.isna(value):
        return None
    if isinstance(value, (pd.Timestamp, datetime, date)):
        return pd.Timestamp(value).isoformat()
    if isinstance(value, (np.integer, int)):
        return int(value)
    if isinstance(value, Decimal):
        if not value.is_finite():
            raise TradeLedgerDataError("Trade identity contains a non-finite number.")
        return format(value, "f")
    if isinstance(value, (np.floating, float)):
        number = float(value)
        if not math.isfinite(number):
            raise TradeLedgerDataError("Trade identity contains a non-finite number.")
        return format(number, ".17g")
    return str(value)


def _trade_key(row: pd.Series) -> str:
    identity = {
        column: _identity_value(row[column])
        for column in IDENTITY_COLUMNS
    }
    payload = json.dumps(identity, sort_keys=True, separators=(",", ":"))
    return hashlib.sha256(payload.encode("utf-8")).hexdigest()


def _json_value(value: Any, *, date_only: bool = False) -> Any:
    if value is None or pd.isna(value):
        return None
    if isinstance(value, (pd.Timestamp, datetime, date)):
        parsed = pd.Timestamp(value)
        return parsed.date().isoformat() if date_only else parsed.isoformat()
    if isinstance(value, (np.integer, int)):
        return int(value)
    if isinstance(value, Decimal):
        return float(value) if value.is_finite() else None
    if isinstance(value, (np.floating, float)):
        number = float(value)
        return number if math.isfinite(number) else None
    if isinstance(value, (np.bool_, bool)):
        return bool(value)
    return str(value)


def build_trade_workbook(
    records: Sequence[Mapping[str, Any]],
    metadata: Mapping[str, Any],
    *,
    columns: Sequence[str],
    labels: Mapping[str, str],
    filter_model: Mapping[str, Any] | None = None,
    selected_substrategies: Iterable[str] | None = None,
) -> bytes:
    """Create a formula-safe workbook matching the current grid row order."""
    workbook = Workbook()
    trades_sheet = workbook.active
    trades_sheet.title = "Trades"
    metadata_sheet = workbook.create_sheet("Snapshot Metadata")

    export_columns = [
        column
        for column in columns
        if column != "_trade_key" and any(column in record for record in records)
    ]
    trades_sheet.append([labels.get(column, column) for column in export_columns])
    for record in records:
        trades_sheet.append(
            [_excel_value(column, record.get(column)) for column in export_columns]
        )

    number_formats = {
        **{
            column: "#,##0.00000000"
            for column in ORIGINAL_MODEL_GREEK_COLUMNS
        },
        **{
            column: "#,##0.000000"
            for column in ORIGINAL_POSITION_GREEK_COLUMNS
        },
        **{
            column: "#,##0.00"
            for column in MONETARY_TOTAL_COLUMNS
        },
    }
    for column_index, column in enumerate(export_columns, start=1):
        number_format = number_formats.get(column)
        if number_format is None:
            continue
        for row_index in range(2, trades_sheet.max_row + 1):
            trades_sheet.cell(
                row=row_index,
                column=column_index,
            ).number_format = number_format

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in trades_sheet[1]:
        cell.font = Font(bold=True, color="1F2937")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    trades_sheet.freeze_panes = "A2"
    trades_sheet.auto_filter.ref = trades_sheet.dimensions
    trades_sheet.sheet_view.showGridLines = False
    for index, column in enumerate(export_columns, start=1):
        header = labels.get(column, column)
        sample_lengths = [
            len(str(trades_sheet.cell(row=row, column=index).value or ""))
            for row in range(2, min(trades_sheet.max_row, 102) + 1)
        ]
        trades_sheet.column_dimensions[
            trades_sheet.cell(row=1, column=index).column_letter
        ].width = min(max([len(header), *sample_lengths], default=len(header)) + 2, 42)

    metadata_rows = [
        ("Source", metadata.get("source")),
        ("COB date", metadata.get("cob_date")),
        ("Valuation run ID", metadata.get("valuation_run_id")),
        ("Valuation revision", metadata.get("valuation_revision")),
        ("Methodology version", metadata.get("valuation_methodology_version")),
        ("Published at", metadata.get("valuation_published_at")),
        ("Published by", metadata.get("valuation_published_by")),
        ("Snapshot row count", metadata.get("row_count")),
        (
            "Native currencies",
            ", ".join(metadata.get("currencies") or []),
        ),
        ("Exported row count", len(records)),
        (
            "Selected substrategies",
            ", ".join(sorted(selected_substrategies or [])) or "All",
        ),
        (
            "Grid filters",
            json.dumps(filter_model or {}, sort_keys=True, default=str),
        ),
        (
            "Premium convention",
            "Signed execution-price basis per native contract currency. Futures-style amounts are not described as universal upfront cashflows.",
        ),
        (
            "Greek convention",
            "Original signed model and quantity Greek fields from the published valuation run.",
        ),
        (
            "Model Greek units",
            "delta/gamma/vega/theta/correlation/rho are the original per-unit model outputs stored at eight decimal places.",
        ),
        (
            "Quantity Greek units",
            "Position Greeks and risk exposures retain the original persisted six decimal places.",
        ),
        (
            "Quantity monetary units",
            "Position value, premium, P&L, intrinsic value, and time value retain the original persisted two decimal places.",
        ),
    ]
    metadata_sheet.append(["Field", "Value"])
    for key, value in metadata_rows:
        metadata_sheet.append([key, _excel_value("", value)])
    for cell in metadata_sheet[1]:
        cell.font = Font(bold=True, color="1F2937")
        cell.fill = header_fill
    metadata_sheet.freeze_panes = "A2"
    metadata_sheet.sheet_view.showGridLines = False
    metadata_sheet.column_dimensions["A"].width = 28
    metadata_sheet.column_dimensions["B"].width = 96

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _excel_value(column: str, value: Any) -> Any:
    if value is None or (isinstance(value, float) and math.isnan(value)):
        return None
    if column in DATE_COLUMNS:
        parsed = pd.to_datetime(value, errors="coerce")
        return None if pd.isna(parsed) else parsed.date()
    if column in TIMESTAMP_COLUMNS:
        parsed = pd.to_datetime(value, errors="coerce")
        if pd.isna(parsed):
            return None
        if parsed.tzinfo is not None:
            parsed = parsed.tz_localize(None)
        return parsed.to_pydatetime()
    if isinstance(value, str):
        return _safe_excel_text(value)
    if isinstance(value, (np.integer, int)):
        return int(value)
    if isinstance(value, Decimal):
        return float(value) if value.is_finite() else None
    if isinstance(value, (np.floating, float)):
        return float(value)
    return _safe_excel_text(value)


def _safe_excel_text(value: Any) -> str | None:
    if value is None:
        return None
    text_value = str(value)
    if text_value.startswith(("=", "+", "-", "@")):
        return "'" + text_value
    return text_value
