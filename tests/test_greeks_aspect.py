import base64
import datetime as dt
import io
from pathlib import Path

import pandas as pd
import pytest
from dash._callback import GLOBAL_CALLBACK_LIST
from openpyxl import load_workbook

from pages import greeks
from runtime_config import clear_runtime_config_cache


@pytest.fixture(autouse=True)
def _reset_runtime_config():
    clear_runtime_config_cache()
    yield
    clear_runtime_config_cache()


def _aspect_frame():
    return pd.DataFrame([
        {
            'instrument': 'ICE_TTF',
            'qty': '10000',
            'uoM': 'MMBtu',
            'maturityForwardDate': '2026-09-01',
            'strategy': 'LNG',
            'dealType': 'Future',
            'entityType': 'Physical',
            'exchangeTradeOtc': 'Exchange',
        },
        {
            'instrument': 'ICE_BRENT_FUTURES',
            'qty': '-2000',
            'uoM': 'BBL',
            'maturityForwardDate': '2026-10-01',
            'strategy': 'LNG',
            'dealType': 'Future',
            'entityType': 'Physical',
            'exchangeTradeOtc': 'Exchange',
        },
    ])


def _aspect_meta():
    return {
        'status': 'ok',
        'message': 'Aspect exposure loaded',
        'source': greeks.ASPECT_SOURCE_LABEL,
        'mode': 'settlement',
        'requested_cob_date': '2026-07-28',
        'fetched_at': '2026-07-28T10:00:00+04:00',
        'source_rows': 2,
        'accepted_rows': 2,
        'rejected_rows': 0,
        'instruments': 2,
        'strategies': 1,
        'book': 'AD-LNG',
    }


def _normalized_greek_row(
    maturity,
    exposure,
    *,
    risk_bucket='ICE_TTF',
    unit='MWh',
    greek='delta',
    bucket_type='Instrument',
    source_row_id=0,
):
    return {
        'cob_date': '2026-07-30',
        'source_row_id': source_row_id,
        'source': 'Options valuation',
        'strategy': 'Test strategy',
        'trade_type': 'Financial Option',
        'option_type': 'vanilla',
        'put_call': 'call',
        'currency': 'USD',
        'greek': greek,
        'greek_label': greeks.GREEK_DEFINITIONS[greek]['label'],
        'bucket_type': bucket_type,
        'risk_bucket': risk_bucket,
        'instrument': risk_bucket if bucket_type == 'Instrument' else 'N/A',
        'unit': unit,
        'maturity_bucket': maturity,
        'maturity_pair': maturity,
        'asset_pair': risk_bucket,
        'exposure': exposure,
        'quantity': 1.0,
        'value': 0.0,
        'pnl': 0.0,
    }


def _walk(component):
    yield component
    children = getattr(component, 'children', None)
    if isinstance(children, (list, tuple)):
        for child in children:
            yield from _walk(child)
    elif children is not None:
        yield from _walk(children)


def test_layout_omits_current_greeks_summary_section():
    components = list(_walk(greeks.layout))
    assert not any(
        getattr(component, 'id', None) == 'greeks-summary-table-container'
        for component in components
    )
    assert not any(
        getattr(component, 'children', None) == 'Current Greeks Summary'
        for component in components
    )
    assert len(greeks.update_monitor_tables(
        greeks._empty_store(),
        {'meta': {'status': 'idle'}, 'rows': []},
        'mixed',
        None,
        None,
        'native',
    )) == 6


def test_corrupt_greeks_snapshot_reference_recovers_as_expired():
    payload = greeks._resolve_greeks_payload('corrupt-session-value')

    assert payload['rows'] == []
    assert payload['meta']['message'] == 'Server snapshot expired; refresh the page'
    assert greeks.export_greeks_workbook(
        1,
        'corrupt-session-value',
        None,
        'mixed',
        None,
        None,
        'native',
    ) is greeks.dash.no_update


def test_common_maturity_axis_aligns_core_and_aspect_tables_with_blank_gaps():
    core_rows = pd.DataFrame([
        _normalized_greek_row('2026-10', 0.0, source_row_id=0),
        _normalized_greek_row(
            '2027-Q1',
            20.0,
            risk_bucket='ICE_HH',
            unit='MMBtu',
            source_row_id=1,
        ),
    ])
    aspect_rows = pd.DataFrame([
        _normalized_greek_row(
            '2028',
            5.0,
            risk_bucket='ICE_BRENT_FUTURES',
            unit='BBL',
            source_row_id='aspect-0',
        ),
    ])

    maturity_axis = greeks._common_maturity_axis(core_rows, aspect_rows)
    assert maturity_axis == ['2026-10', '2027-Q1', '2028']

    tables = greeks.build_display_tables(core_rows, maturity_axis)
    expected_rows = maturity_axis + ['Total']
    for bucket_table in tables['bucket_greek_tables']:
        assert bucket_table['table']['Maturity'].tolist() == expected_rows
    assert tables['delta_ladder']['Maturity'].tolist() == expected_rows
    assert tables['delta_unit_ladder']['Maturity'].tolist() == expected_rows

    aspect_delta = greeks.create_ladder_df(
        aspect_rows,
        'delta',
        maturity_axis,
    )
    assert aspect_delta['Maturity'].tolist() == expected_rows

    ttf_table = next(
        item['table']
        for item in tables['bucket_greek_tables']
        if item['risk_bucket'] == 'TTF'
    ).set_index('Maturity')
    assert ttf_table.at['2026-10', 'Delta'] == 0.0
    assert pd.isna(ttf_table.at['2027-Q1', 'Delta'])
    assert pd.isna(ttf_table.at['2028', 'Delta'])
    assert ttf_table.at['Total', 'Delta'] == 0.0


def test_common_maturity_axis_preserves_business_labels_and_sort_order():
    rows = pd.DataFrame({
        'maturity_bucket': [
            'Unknown',
            '2028',
            '2027-CAL',
            '2027-Q2',
            '2027-03 / 2027-04',
        ],
    })

    assert greeks._common_maturity_axis(rows) == [
        '2027-03 / 2027-04',
        '2027-Q2',
        '2027-CAL',
        '2028',
        'Unknown',
    ]


def test_grid_records_distinguish_real_zero_from_missing_data():
    frame = pd.DataFrame([
        {'Maturity': '2026-10', 'Delta': 0.0},
        {'Maturity': '2026-11', 'Delta': None},
    ])

    records = greeks._format_grid_records(frame, ['Delta'])

    assert records[0]['Delta'] == '0'
    assert records[0]['__raw_Delta'] == 0.0
    assert records[1]['Delta'] is None
    assert '__raw_Delta' not in records[1]


def test_comparison_grid_locks_chronology_and_enables_local_scrolling():
    grid = greeks.build_compact_table(
        'comparison-grid-test',
        pd.DataFrame([
            {'Maturity': '2026-10', 'Delta': 1.0, '_row_type': 'normal'},
            {'Maturity': 'Total', 'Delta': 1.0, '_row_type': 'total'},
        ]),
        comparison_grid=True,
        accessible_label='Delta by maturity and asset',
    )

    assert 'greeks-comparison-grid' in grid.className.split()
    assert grid.dashGridOptions['rowHeight'] == 28
    assert grid.dashGridOptions['headerHeight'] == 30
    assert grid.dashGridOptions['groupHeaderHeight'] == 24
    assert grid.dashGridOptions['suppressHorizontalScroll'] is False
    assert grid.dashGridOptions['suppressMovableColumns'] is True
    assert grid.dashGridOptions['getRowId'] == {
        'function': "params.data['Maturity']",
    }
    assert grid.dashGridOptions['ariaLabel'] == 'Delta by maturity and asset'
    assert grid.columnSize is None

    maturity_column = next(
        column for column in grid.columnDefs if column.get('field') == 'Maturity'
    )
    delta_column = next(
        column for column in grid.columnDefs if column.get('field') == 'Delta'
    )
    assert maturity_column['pinned'] == 'left'
    assert maturity_column['lockPinned'] is True
    assert maturity_column['sortable'] is False
    assert maturity_column['resizable'] is False
    assert maturity_column['width'] == 68
    assert maturity_column['minWidth'] == maturity_column['width']
    assert maturity_column['maxWidth'] == maturity_column['width']
    assert delta_column['sortable'] is False
    assert delta_column['resizable'] is False
    assert delta_column['minWidth'] == delta_column['width']
    assert delta_column['maxWidth'] == delta_column['width']
    assert grid.style['width'] == (
        f"{maturity_column['width'] + delta_column['width'] + 2}px"
    )


def test_comparison_columns_expand_to_fit_the_longest_header_or_value():
    short_grid = greeks.build_compact_table(
        'comparison-short-grid-test',
        pd.DataFrame([{'Maturity': '2026-10', 'Delta': 1.0}]),
        comparison_grid=True,
    )
    long_grid = greeks.build_compact_table(
        'comparison-long-grid-test',
        pd.DataFrame([{'Maturity': '2027-CAL', 'Delta': -123456789.0}]),
        comparison_grid=True,
    )

    short_delta = next(
        column for column in short_grid.columnDefs if column.get('field') == 'Delta'
    )
    long_delta = next(
        column for column in long_grid.columnDefs if column.get('field') == 'Delta'
    )
    assert long_delta['width'] > short_delta['width']
    assert 'flex' not in long_delta


def test_comparison_table_css_packs_tables_and_uses_dark_headers():
    css = (
        Path(__file__).resolve().parents[1] / 'assets' / 'styles.css'
    ).read_text(encoding='utf-8')

    assert '.greeks-bucket-greek-grid-wrap,' in css
    assert 'display: flex;\n    flex-wrap: wrap;' in css
    assert '--ag-header-background-color: #1e293b;' in css
    assert '--ag-header-foreground-color: #ffffff;' in css
    assert 'color: #ffffff !important;' in css


def test_aggregation_and_unit_use_compact_segmented_controls():
    components = list(_walk(greeks.layout))
    aggregation = next(
        component
        for component in components
        if getattr(component, 'id', None) == 'maturity-aggregation-mode-selector'
    )
    unit = next(
        component
        for component in components
        if getattr(component, 'id', None) == 'unit-mode-selector'
    )

    assert aggregation.value == 'mixed'
    assert unit.value == 'native'
    assert aggregation.inline is True
    assert unit.inline is True
    assert 'greeks-segmented-control' in aggregation.className.split()
    assert 'greeks-segmented-control' in unit.className.split()
    assert getattr(aggregation, 'inputStyle', None) in (None, {})
    assert getattr(unit, 'inputStyle', None) in (None, {})
    assert any(
        getattr(component, 'children', None) == 'Unit'
        and 'inline-filter-label' in getattr(component, 'className', '').split()
        for component in components
    )
    assert not any(
        getattr(component, 'children', None) == 'Unit Mode'
        for component in components
    )


def test_maturity_cutoff_labels_are_complete_and_compact():
    assert greeks._month_through_label('2027-03-31') == '2027 Q1 · Mar'
    assert greeks._month_through_label('2027-12-31') == '2027 Q4 · Dec'
    assert greeks._quarter_through_label('2027-12-31') == '2027'


def test_all_selectors_and_actions_share_one_toolbar_row():
    selector_row = next(
        component
        for component in _walk(greeks.layout)
        if 'greeks-monitor-selector-row'
        in getattr(component, 'className', '').split()
    )
    selector_ids = [
        next(
            getattr(descendant, 'id', None)
            for descendant in _walk(control_group)
            if getattr(descendant, 'id', None)
        )
        for control_group in selector_row.children[:8]
    ]
    action_ids = {
        getattr(component, 'id', None)
        for component in _walk(selector_row.children[8])
        if getattr(component, 'id', None)
    }

    assert selector_ids == [
        'date-selector',
        'maturity-aggregation-mode-selector',
        'month-through-selector',
        'quarter-through-selector',
        'unit-mode-selector',
        'strategy-selector',
        'trade-type-selector',
        'risk-bucket-selector',
    ]
    assert action_ids == {
        'greeks-aspect-settlement-btn',
        'greeks-aspect-live-btn',
        'export-greeks-workbook-btn',
    }
    assert len(selector_row.children) == 10
    assert 'greeks-control-actions' in selector_row.children[8].className.split()
    assert selector_row.children[8].children[0].children == 'Aspect Actions'
    assert 'greeks-inline-source-status' in selector_row.children[9].className.split()
    inline_source_ids = {
        getattr(component, 'id', None)
        for component in _walk(selector_row.children[9])
        if getattr(component, 'id', None)
    }
    assert inline_source_ids == {
        'greeks-source-status-mount',
        'greeks-source-status-inline',
    }
    assert not any(
        'greeks-monitor-action-row' in getattr(component, 'className', '').split()
        for component in _walk(greeks.layout)
    )

    action_buttons = {
        getattr(component, 'id', None): component
        for component in _walk(selector_row.children[8])
        if getattr(component, 'id', None)
    }
    assert action_buttons['greeks-aspect-settlement-btn'].children == 'COB'
    assert getattr(
        action_buttons['greeks-aspect-settlement-btn'],
        'aria-label',
    ) == 'Load Aspect COB'
    assert action_buttons['greeks-aspect-live-btn'].children == 'Live'
    assert getattr(
        action_buttons['greeks-aspect-live-btn'],
        'aria-label',
    ) == 'Load Aspect Live'
    assert action_buttons['export-greeks-workbook-btn'].children == 'Export'
    assert getattr(
        action_buttons['export-greeks-workbook-btn'],
        'aria-label',
    ) == 'Export Workbook'


def test_toolbar_css_reserves_dropdown_icon_space_and_collapses_before_squeezing():
    css = (
        Path(__file__).resolve().parents[1] / 'assets' / 'styles.css'
    ).read_text(encoding='utf-8')

    assert '100px\n        84px\n        198px' in css
    assert '194px\n        159px\n        124px\n        minmax(260px, 1fr);' in css
    assert (
        'grid-template-columns: minmax(0, 1fr) 14px 16px !important;'
        in css
    )
    assert '@media (max-width: 1329px)' in css
    assert '@media (max-width: 1573px) and (min-width: 1330px)' in css
    assert '124px\n            16px;' in css
    assert 'grid-template-columns: repeat(5, minmax(0, 1fr));' in css
    assert '@media (max-width: 900px)' in css
    assert 'grid-template-columns: repeat(3, minmax(0, 1fr));' in css


def test_strategy_filter_uses_compact_debounced_multi_picker():
    strategy_filter = next(
        component
        for component in _walk(greeks.layout)
        if getattr(component, 'id', None) == 'strategy-selector'
    )

    assert strategy_filter.multi is True
    assert strategy_filter.closeOnSelect is False
    assert strategy_filter.debounce is True
    assert strategy_filter.optionHeight == 40
    assert strategy_filter.maxHeight == 360
    assert strategy_filter.labels['selected_count'] == '{num_selected} strategies selected'
    assert strategy_filter.labels['search'] == 'Search strategies'
    assert 'greeks-compact-multi-dropdown' in strategy_filter.className.split()
    assert 'greeks-strategy-dropdown' in strategy_filter.className.split()
    assert 'All strategies selected' in greeks.STRATEGY_DROPDOWN_LABELS_CLIENTSIDE
    assert 'availableValues.every' in greeks.STRATEGY_DROPDOWN_LABELS_CLIENTSIDE

    label_callback = next(
        item
        for item in GLOBAL_CALLBACK_LIST
        if item.get('output') == 'strategy-selector.labels'
    )
    assert label_callback['inputs'] == [
        {'id': 'strategy-selector', 'property': 'options'},
        {'id': 'strategy-selector', 'property': 'value'},
    ]
    assert label_callback['clientside_function']


def test_asset_pair_filter_uses_compact_debounced_multi_picker():
    risk_filter = next(
        component
        for component in _walk(greeks.layout)
        if getattr(component, 'id', None) == 'risk-bucket-selector'
    )

    assert risk_filter.multi is True
    assert risk_filter.closeOnSelect is False
    assert risk_filter.debounce is True
    assert risk_filter.optionHeight == 40
    assert risk_filter.maxHeight == 360
    assert risk_filter.labels['selected_count'] == '{num_selected} assets / pairs'
    assert risk_filter.labels['search'] == 'Search assets and pairs'
    assert 'greeks-compact-multi-dropdown' in risk_filter.className.split()
    assert 'greeks-risk-bucket-dropdown' in risk_filter.className.split()
    assert 'All assets / pairs' in greeks.RISK_BUCKET_DROPDOWN_LABELS_CLIENTSIDE
    assert 'availableValues.every' in greeks.RISK_BUCKET_DROPDOWN_LABELS_CLIENTSIDE

    label_callback = next(
        item
        for item in GLOBAL_CALLBACK_LIST
        if item.get('output') == 'risk-bucket-selector.labels'
    )
    assert label_callback['inputs'] == [
        {'id': 'risk-bucket-selector', 'property': 'options'},
        {'id': 'risk-bucket-selector', 'property': 'value'},
    ]
    assert label_callback['clientside_function']


def test_trade_type_filter_uses_compact_debounced_multi_picker():
    trade_type_filter = next(
        component
        for component in _walk(greeks.layout)
        if getattr(component, 'id', None) == 'trade-type-selector'
    )

    assert trade_type_filter.multi is True
    assert trade_type_filter.closeOnSelect is False
    assert trade_type_filter.debounce is True
    assert trade_type_filter.optionHeight == 40
    assert trade_type_filter.maxHeight == 360
    assert trade_type_filter.labels['selected_count'] == '{num_selected} trade types selected'
    assert trade_type_filter.labels['search'] == 'Search trade types'
    assert 'greeks-compact-multi-dropdown' in trade_type_filter.className.split()
    assert 'greeks-trade-type-dropdown' in trade_type_filter.className.split()
    assert 'All trade types selected' in greeks.TRADE_TYPE_DROPDOWN_LABELS_CLIENTSIDE
    assert 'availableValues.every' in greeks.TRADE_TYPE_DROPDOWN_LABELS_CLIENTSIDE

    label_callback = next(
        item
        for item in GLOBAL_CALLBACK_LIST
        if item.get('output') == 'trade-type-selector.labels'
    )
    assert label_callback['inputs'] == [
        {'id': 'trade-type-selector', 'property': 'options'},
        {'id': 'trade-type-selector', 'property': 'value'},
    ]
    assert label_callback['clientside_function']


def test_option_assets_use_contract_native_units(monkeypatch):
    monkeypatch.setattr(
        greeks,
        'fetch_instrument_unit_map',
        lambda: {
            'ICE_TTF': {
                'native_unit': 'MWh',
                'unit': 'MMBtu',
                'conv_factor': 3.41214245,
            },
        },
    )
    data = pd.DataFrame([
        {
            'cob_date': '2026-07-27',
            'unit_quantity': 'MWh',
            'asset_a': 'ICE_TTF',
            'asset_b': None,
            'maturity_date_a': '2026-10-01',
            'maturity_date_type_a': 'month',
            'quantity': 74500,
            'qty_delta_asset_a': 100.0,
            'qty_gamma_asset_a': 10.0,
            'qty_vega_sigma1': 20.0,
            'qty_theta': 5.0,
            'qty_corr_sensitivity': 0.0,
        },
    ])

    normalized = greeks.normalize_greek_contributions(
        data,
        aggregation='month',
        unit_mode='native',
        cob_date='2026-07-27',
    )
    asset_rows = normalized[normalized['bucket_type'] == 'Instrument']

    assert set(asset_rows['unit']) == {'MWh'}
    assert asset_rows.loc[asset_rows['greek'] == 'delta', 'exposure'].iloc[0] == 100.0
    assert asset_rows.loc[asset_rows['greek'] == 'theta', 'exposure'].iloc[0] == 5.0
    assert normalized.loc[normalized['bucket_type'] == 'Pair'].empty

    bucket_tables = greeks.create_bucket_greek_tables(normalized)
    titles = [table['title'] for table in bucket_tables]
    assert 'ASSET: TTF (MWh)' in titles
    assert 'PAIR: TTF (MWh)' not in titles
    ttf_table = next(table for table in bucket_tables if table['title'] == 'ASSET: TTF (MWh)')
    assert list(ttf_table['table'].columns) == [
        'Maturity',
        'Delta',
        'Gamma',
        'Vega',
        'Theta',
        '_row_type',
    ]
    exported = greeks.create_bucket_greek_export_df(bucket_tables)
    assert set(exported['Bucket Type']) == {'Instrument'}
    assert set(exported['Risk Bucket']) == {'TTF'}
    assert exported.loc[exported['Maturity'] == 'Total', 'Theta'].iloc[0] == 5.0


def test_spread_theta_and_correlation_remain_pair_risks(monkeypatch):
    monkeypatch.setattr(
        greeks,
        'fetch_instrument_unit_map',
        lambda: {
            'ICE_HH': {
                'native_unit': 'MMBtu',
                'unit': 'MMBtu',
                'conv_factor': 1.0,
            },
            'ICE_JKM': {
                'native_unit': 'MMBtu',
                'unit': 'MMBtu',
                'conv_factor': 1.0,
            },
        },
    )
    data = pd.DataFrame([
        {
            'cob_date': '2026-07-27',
            'unit_quantity': 'MMBtu',
            'asset_a': 'ICE_HH',
            'asset_b': 'ICE_JKM',
            'maturity_date_a': '2027-02-01',
            'maturity_date_type_a': 'month',
            'maturity_date_b': '2027-02-01',
            'maturity_date_type_b': 'month',
            'quantity': 10000,
            'qty_delta_asset_a': 100.0,
            'qty_delta_asset_b': -80.0,
            'qty_gamma_asset_a': 10.0,
            'qty_gamma_asset_b': 8.0,
            'qty_vega_sigma1': 20.0,
            'qty_vega_sigma2': 15.0,
            'qty_theta': -12.0,
            'qty_corr_sensitivity': 4.0,
        },
    ])

    normalized = greeks.normalize_greek_contributions(
        data,
        aggregation='month',
        unit_mode='native',
        cob_date='2026-07-27',
    )
    pair_rows = normalized[normalized['bucket_type'] == 'Pair']
    assert set(pair_rows['greek']) == {'theta', 'correlation'}
    assert set(pair_rows['risk_bucket']) == {'ICE_HH / ICE_JKM'}

    bucket_tables = greeks.create_bucket_greek_tables(normalized)
    pair_table = next(table for table in bucket_tables if table['bucket_type'] == 'Pair')
    assert list(pair_table['table'].columns) == [
        'Maturity',
        'Theta',
        'Correlation',
        '_row_type',
    ]
    for asset_table in (
        table for table in bucket_tables if table['bucket_type'] == 'Instrument'
    ):
        assert 'Theta' not in asset_table['table'].columns


def test_greek_normalization_uses_stable_source_ids_for_non_default_index(monkeypatch):
    monkeypatch.setattr(
        greeks,
        'fetch_instrument_unit_map',
        lambda: {
            'ICE_TTF': {
                'native_unit': 'MWh',
                'unit': 'MMBtu',
                'conv_factor': 3.41214245,
            },
        },
    )
    data = pd.DataFrame(
        [
            {
                'cob_date': pd.Timestamp('2026-07-27'),
                'unit_quantity': 'MWh',
                'asset_a': 'ICE_TTF',
                'asset_b': None,
                'maturity_date_a': pd.Timestamp('2026-10-01'),
                'maturity_date_type_a': 'month',
                'qty_delta_asset_a': 100.0,
                'qty_gamma_asset_a': 10.0,
                'qty_vega_sigma1': 20.0,
                'qty_theta': 5.0,
            }
        ],
        index=[42],
    )

    normalized = greeks.normalize_greek_contributions(
        data,
        aggregation='mixed',
        unit_mode='native',
        cob_date='2026-07-27',
    )

    assert normalized['source_row_id'].unique().tolist() == [0]
    assert normalized['cob_date'].unique().tolist() == ['2026-07-27']
    assert normalized['maturity_bucket'].unique().tolist() == ['2026-10']


def test_filter_values_only_labels_two_underlying_rows_as_pairs(monkeypatch):
    data = pd.DataFrame([
        {
            'substrategy': 'TTF call spread',
            'type_trade': 'Financial Option',
            'asset_a': 'ICE_TTF',
            'asset_b': None,
        },
        {
            'substrategy': 'Brent put fly',
            'type_trade': 'Financial Option',
            'asset_a': 'ICE_BRENT_FUTURES',
            'asset_b': None,
        },
        {
            'substrategy': 'JKM-HH spread',
            'type_trade': 'Physical Option',
            'asset_a': 'ICE_HH',
            'asset_b': 'ICE_JKM',
        },
    ])
    monkeypatch.setattr(greeks, 'fetch_options_data', lambda _cob_date: data)

    _, _, bucket_options = greeks.fetch_filter_values('2026-07-27')
    values = {option['value'] for option in bucket_options}

    assert 'pair::ICE_TTF' not in values
    assert 'pair::ICE_BRENT_FUTURES' not in values
    assert 'pair::ICE_HH / ICE_JKM' in values


@pytest.mark.parametrize(
    ('maturity_type_a', 'maturity_type_b'),
    [
        ('month', 'month'),
        ('calendar', 'calendar'),
        ('calendar', 'month'),
    ],
)
def test_mixed_annual_bucket_combines_all_delivery_structures(
    maturity_type_a,
    maturity_type_b,
):
    row = {
        'maturity_date_a': '2029-01-01',
        'maturity_date_type_a': maturity_type_a,
        'maturity_date_b': '2029-01-01',
        'maturity_date_type_b': maturity_type_b,
    }

    assert greeks._format_maturity_pair(
        row,
        aggregation='mixed',
        month_through='2027-03-31',
        quarter_through='2027-12-31',
        cob_date='2026-07-27',
    ) == '2029'


def test_calendar_marker_is_preserved_outside_annual_aggregation():
    row = {
        'maturity_date_a': '2029-01-01',
        'maturity_date_type_a': 'calendar',
        'maturity_date_b': '2029-01-01',
        'maturity_date_type_b': 'month',
    }

    assert greeks._format_maturity_pair(
        row,
        aggregation='quarter',
        cob_date='2026-07-27',
    ) == '2029-CAL / 2029-Q1'


def test_aspect_request_uses_json_auth_timeout_and_ssl_verification(
    monkeypatch,
    tmp_path,
):
    config_path = tmp_path / 'empty.ini'
    config_path.write_text('')
    monkeypatch.setenv('OPTIONS_CONFIG_PATH', str(config_path))
    monkeypatch.setenv('ASPECT_USERNAME', 'user')
    monkeypatch.setenv('ASPECT_PASSWORD', 'password')
    monkeypatch.setenv('ASPECT_REQUEST_TIMEOUT_SECONDS', '45')
    monkeypatch.delenv('ASPECT_VERIFY_SSL', raising=False)
    clear_runtime_config_cache()

    captured = {}

    class _Response:
        status_code = 200
        ok = True

        @staticmethod
        def json():
            return _aspect_frame().to_dict('records')

    def fake_post(url, **kwargs):
        captured['url'] = url
        captured.update(kwargs)
        return _Response()

    monkeypatch.setattr(greeks.requests, 'post', fake_post)
    monkeypatch.setattr(greeks, '_dubai_today', lambda: dt.date(2026, 7, 28))

    result = greeks.fetch_aspect_exposure_report(dt.date(2026, 7, 27))

    assert len(result) == 2
    assert captured['url'] == greeks.ASPECT_DEFAULT_URL
    assert captured['json'] == {
        'cobDate': '2026-07-27',
        'todayPricing': 1,
        'book': 'AD-LNG',
        'displayExchangeTradeOtc': True,
    }
    assert captured['auth'] == ('user', 'password')
    assert captured['timeout'] == (10.0, 45.0)
    assert captured['verify'] is True
    assert 'proxies' not in captured


def test_missing_aspect_credentials_fails_closed(monkeypatch, tmp_path):
    config_path = tmp_path / 'empty.ini'
    config_path.write_text('')
    monkeypatch.setenv('OPTIONS_CONFIG_PATH', str(config_path))
    monkeypatch.delenv('ASPECT_USERNAME', raising=False)
    monkeypatch.delenv('ASPECT_PASSWORD', raising=False)
    clear_runtime_config_cache()

    with pytest.raises(greeks.AspectConfigurationError, match='credentials are unavailable'):
        greeks.fetch_aspect_exposure_report(dt.date(2026, 7, 28))


def test_prepare_aspect_records_rejects_invalid_rows():
    data = _aspect_frame()
    invalid = data.iloc[[0]].copy()
    invalid['instrument'] = ''
    invalid['qty'] = 'not-a-number'
    invalid['maturityForwardDate'] = 'not-a-date'
    data = pd.concat([data, invalid], ignore_index=True)

    prepared, rejected_rows = greeks._prepare_aspect_records(data)

    assert len(prepared) == 2
    assert rejected_rows == 1
    assert prepared['qty'].tolist() == [10000.0, -2000.0]
    assert prepared['maturityForwardDate'].tolist() == ['2026-09-01', '2026-10-01']


def test_aspect_normalization_applies_mapping_and_lot_conventions(monkeypatch):
    prepared, _ = greeks._prepare_aspect_records(_aspect_frame())
    monkeypatch.setattr(
        greeks,
        'fetch_instrument_unit_map',
        lambda: {
            'ICE_TTF': {'unit': 'MMBtu', 'conv_factor': 1.0},
            'ICE_BRENT_FUTURES': {'unit': 'BBL', 'conv_factor': 1.0},
        },
    )

    normalized = greeks.normalize_aspect_contributions(
        prepared,
        aggregation='month',
        unit_mode='lots',
        cob_date='2026-07-28',
    )

    exposure_by_instrument = normalized.set_index('risk_bucket')['exposure'].to_dict()
    assert exposure_by_instrument == {
        'ICE_TTF': 1.0,
        'ICE_BRENT_FUTURES': -2.0,
    }
    assert set(normalized['greek']) == {'delta'}
    assert set(normalized['source']) == {greeks.ASPECT_SOURCE_LABEL}


def test_aspect_only_export_contains_source_sheets(monkeypatch):
    prepared, _ = greeks._prepare_aspect_records(_aspect_frame())
    monkeypatch.setattr(
        greeks,
        'fetch_instrument_unit_map',
        lambda: {
            'ICE_TTF': {'unit': 'MMBtu', 'conv_factor': 1.0},
            'ICE_BRENT_FUTURES': {'unit': 'BBL', 'conv_factor': 1.0},
        },
    )

    result = greeks.export_greeks_workbook(
        1,
        greeks._empty_store(),
        {'meta': _aspect_meta(), 'rows': prepared.to_dict('records')},
        'month',
        None,
        None,
        'lots',
    )
    workbook = load_workbook(
        io.BytesIO(base64.b64decode(result['content'])),
        read_only=True,
        data_only=False,
    )

    assert workbook.sheetnames == [
        'Aspect Metadata',
        'Aspect Summary',
        'Aspect Delta',
        'Aspect Unit Delta',
        'Aspect Raw',
    ]
    assert workbook['Aspect Metadata']['B3'].value == 'settlement'
    assert workbook['Aspect Raw']['A2'].value == 'ICE_TTF'
    aspect_summary = workbook['Aspect Summary']
    summary_headers = [
        cell.value for cell in next(aspect_summary.iter_rows(min_row=1, max_row=1))
    ]
    delta_column = summary_headers.index('Delta') + 1
    assert (
        aspect_summary.cell(row=2, column=delta_column).number_format
        == '#,##0.000000'
    )


def test_export_uses_shared_maturity_axis_and_preserves_blank_vs_zero(monkeypatch):
    prepared, _ = greeks._prepare_aspect_records(_aspect_frame())
    monkeypatch.setattr(
        greeks,
        'fetch_instrument_unit_map',
        lambda: {
            'ICE_TTF': {
                'native_unit': 'MWh',
                'unit': 'MMBtu',
                'conv_factor': 1.0,
            },
            'ICE_BRENT_FUTURES': {
                'native_unit': 'BBL',
                'unit': 'BBL',
                'conv_factor': 1.0,
            },
        },
    )
    core_rows = pd.DataFrame([
        _normalized_greek_row('2026-10', 0.0),
    ])
    core_store = {
        'meta': {'cob_date': '2026-07-30'},
        'rows': core_rows.to_dict('records'),
    }

    result = greeks.export_greeks_workbook(
        1,
        core_store,
        {'meta': _aspect_meta(), 'rows': prepared.to_dict('records')},
        'month',
        None,
        None,
        'native',
    )
    workbook = load_workbook(
        io.BytesIO(base64.b64decode(result['content'])),
        read_only=True,
        data_only=False,
    )

    bucket_rows = list(workbook['Bucket Greeks'].iter_rows(values_only=True))
    bucket_headers = list(bucket_rows[0])
    maturity_index = bucket_headers.index('Maturity')
    delta_index = bucket_headers.index('Delta')
    displayed = {
        row[maturity_index]: row[delta_index]
        for row in bucket_rows[1:]
    }
    assert list(displayed) == ['2026-09', '2026-10', 'Total']
    assert displayed['2026-09'] is None
    assert displayed['2026-10'] == 0
    assert displayed['Total'] == 0

    aspect_delta_rows = list(
        workbook['Aspect Delta'].iter_rows(values_only=True)
    )
    aspect_maturities = [row[0] for row in aspect_delta_rows[1:]]
    assert aspect_maturities == ['2026-09', '2026-10', 'Total']


def test_quantity_greek_displays_show_no_decimal_places():
    frame = pd.DataFrame([{'Delta': 1234.1234567}])
    rounded = greeks._round_numeric(frame)

    assert rounded.loc[0, 'Delta'] == pytest.approx(1234)
    assert greeks._format_grid_number(rounded.loc[0, 'Delta']) == '1,234'
    assert greeks._format_currency(-12.3456784) == '-$12'
