from __future__ import annotations

from datetime import date, datetime
from decimal import Decimal
from io import BytesIO

import pandas as pd
import pytest
from openpyxl import load_workbook

import trade_ledger
from pages import trades


def _source_row(**overrides):
    row = {}
    for column in trade_ledger.OUTPUT_COLUMNS:
        if column in trade_ledger.DATE_COLUMNS:
            row[column] = date(2026, 7, 27)
        elif column in trade_ledger.TIMESTAMP_COLUMNS:
            row[column] = datetime(2026, 7, 28, 8, 30)
        elif column in trade_ledger.TEXT_COLUMNS:
            row[column] = f"{column}-value"
        elif column == "valuation_run_id":
            row[column] = "run-1"
        else:
            row[column] = 1.0

    row.update(
        {
            "cob_date": date(2026, 7, 27),
            "trade_date": date(2026, 1, 15),
            "expiration_date": date(2026, 12, 18),
            "entity": "LNG",
            "book": "OPTIONS",
            "strategy": "Options",
            "substrategy": "JKM call",
            "type_trade": "option",
            "type_option": "European",
            "model": "Black-76",
            "put_call": "call",
            "buy_sell": "buy",
            "currency": "USD",
            "quantity": 1250.5,
            "unit_quantity": "MMBtu",
            "strike": 12.75,
            "premium": -0.42,
            "qty_premium": -525.21,
            "valuation_run_id": "run-1",
            "valuation_revision": 3,
            "valuation_methodology_version": "portfolio-v3",
            "valuation_input_fingerprint": "fingerprint-1",
            "valuation_created_by": "valuation-job",
            "valuation_published_by": "approver",
        }
    )
    row.update(overrides)
    return row


def _frame(*rows):
    return pd.DataFrame(rows, columns=trade_ledger.OUTPUT_COLUMNS)


def _component_tree(component):
    if isinstance(component, (list, tuple)):
        for child in component:
            yield from _component_tree(child)
        return
    yield component
    children = getattr(component, "children", None)
    if children is None:
        return
    if not isinstance(children, (list, tuple)):
        children = [children]
    for child in children:
        yield from _component_tree(child)


def _column_fields(column_defs):
    for definition in column_defs:
        if "children" in definition:
            yield from _column_fields(definition["children"])
        elif "field" in definition:
            yield definition["field"], definition


def test_snapshot_preserves_every_source_row_and_signed_premium():
    frame = _frame(
        _source_row(),
        _source_row(
            substrategy="TTF put",
            put_call="put",
            strike=9.5,
            premium=0.31,
            qty_premium=387.655,
        ),
    )

    snapshot = trade_ledger._validate_snapshot(frame, date(2026, 7, 27))

    assert snapshot.row_count == len(frame)
    assert snapshot.rows["qty_premium"].tolist() == [-525.21, 387.655]
    assert snapshot.rows["quantity"].dtype.kind == "f"
    assert snapshot.rows["delta_s1"].dtype.kind == "f"
    assert snapshot.rows["qty_delta_asset_a"].dtype.kind == "f"
    assert snapshot.rows["_trade_key"].is_unique
    assert snapshot.records()[0]["trade_date"] == "2026-01-15"
    assert "T" in snapshot.records()[0]["valuation_published_at"]


def test_snapshot_default_order_is_trade_date_then_substrategy_then_expiry():
    snapshot = trade_ledger._validate_snapshot(
        _frame(
            _source_row(
                trade_date=date(2025, 12, 1),
                substrategy="Z strategy",
                strike=11.0,
            ),
            _source_row(
                trade_date=date(2026, 2, 1),
                substrategy="B strategy",
                strike=13.0,
            ),
            _source_row(
                trade_date=date(2026, 2, 1),
                substrategy="A strategy",
                strike=12.0,
            ),
        ),
        date(2026, 7, 27),
    )

    assert snapshot.rows["substrategy"].tolist() == [
        "A strategy",
        "B strategy",
        "Z strategy",
    ]


def test_snapshot_rejects_duplicate_booking_identity():
    with pytest.raises(
        trade_ledger.TradeLedgerDataError,
        match="unique booking identity",
    ):
        trade_ledger._validate_snapshot(
            _frame(_source_row(), _source_row()),
            date(2026, 7, 27),
        )


@pytest.mark.parametrize(
    ("override", "message"),
    [
        ({"valuation_run_id": "run-2"}, "valuation_run_id"),
        ({"valuation_revision": 4}, "valuation_revision"),
        (
            {"valuation_methodology_version": "portfolio-v4"},
            "valuation_methodology_version",
        ),
    ],
)
def test_snapshot_rejects_mixed_run_metadata(override, message):
    with pytest.raises(trade_ledger.TradeLedgerDataError, match=message):
        trade_ledger._validate_snapshot(
            _frame(
                _source_row(),
                _source_row(strike=13.0, **override),
            ),
            date(2026, 7, 27),
        )


@pytest.mark.parametrize(
    ("column", "value"),
    [
        ("delta_s1", None),
        ("gamma_s2", float("inf")),
        ("qty_rho", float("nan")),
    ],
)
def test_snapshot_rejects_incomplete_original_greeks(column, value):
    with pytest.raises(
        trade_ledger.TradeLedgerDataError,
        match="incomplete original valuation Greeks",
    ):
        trade_ledger._validate_snapshot(
            _frame(_source_row(**{column: value})),
            date(2026, 7, 27),
        )


def test_loader_uses_explicit_row_preserving_query_and_strategy_filter(monkeypatch):
    captured = {}
    expected = _frame(_source_row())
    monkeypatch.setattr(trade_ledger, "assert_source_schema", lambda engine: None)

    def fake_read_sql(query, engine, params=None):
        captured["query"] = str(query)
        captured["params"] = params
        return expected

    monkeypatch.setattr(trade_ledger.pd, "read_sql", fake_read_sql)

    snapshot = trade_ledger.load_trade_snapshot(
        object(),
        "2026-07-27",
        ["JKM call"],
    )

    assert snapshot.row_count == 1
    assert "GROUP BY" not in captured["query"].upper()
    assert "substrategy IN" in captured["query"]
    assert "vol_swaption_discount" not in captured["query"]
    for field in (
        "forward_source_name",
        "forward_source_cob_date",
        "volatility_source_name",
        "volatility_source_cob_date",
        "volatility_method",
    ):
        assert field not in captured["query"]
    assert captured["params"]["substrategies"] == ("JKM call",)


def test_loader_surfaces_empty_and_database_error_states(monkeypatch):
    monkeypatch.setattr(trade_ledger, "assert_source_schema", lambda engine: None)
    monkeypatch.setattr(
        trade_ledger.pd,
        "read_sql",
        lambda *args, **kwargs: pd.DataFrame(columns=trade_ledger.OUTPUT_COLUMNS),
    )
    with pytest.raises(trade_ledger.TradeLedgerDataError, match="No active valued"):
        trade_ledger.load_trade_snapshot(object(), "2026-07-27")

    def fail_read(*args, **kwargs):
        raise RuntimeError("database offline")

    monkeypatch.setattr(trade_ledger.pd, "read_sql", fail_read)
    with pytest.raises(RuntimeError, match="database offline"):
        trade_ledger.load_trade_snapshot(object(), "2026-07-27")


def test_cob_discovery_requires_original_source_schema(monkeypatch):
    monkeypatch.setattr(
        trade_ledger,
        "_available_columns",
        lambda engine: {"cob_date"},
    )
    with pytest.raises(
        trade_ledger.TradeLedgerDataError,
        match="schema is incomplete",
    ):
        trade_ledger.get_available_cob_dates(object())


def test_cob_discovery_gates_on_complete_original_greeks_and_single_run(
    monkeypatch,
):
    captured = {}
    monkeypatch.setattr(trade_ledger, "assert_source_schema", lambda engine: None)

    def fake_read_sql(query, engine):
        captured["query"] = str(query)
        return pd.DataFrame({"cob_date": [date(2026, 7, 27)]})

    monkeypatch.setattr(trade_ledger.pd, "read_sql", fake_read_sql)

    assert trade_ledger.get_available_cob_dates(object()) == ["2026-07-27"]
    for column in trade_ledger.ORIGINAL_GREEK_SOURCE_COLUMNS:
        assert f"{column} IS NOT NULL" in captured["query"]
    assert "count(DISTINCT valuation_run_id) = 1" in captured["query"]
    assert "risk_currency" not in captured["query"]


def test_workbook_reopens_with_native_values_order_and_snapshot_metadata():
    snapshot = trade_ledger._validate_snapshot(
        _frame(
            _source_row(substrategy="=unsafe formula", strike=10.0),
            _source_row(substrategy="Safe strategy", strike=11.0),
        ),
        date(2026, 7, 27),
    )
    records = list(reversed(snapshot.records()))
    payload = trade_ledger.build_trade_workbook(
        records,
        snapshot.metadata(),
            columns=[
                "trade_date",
                "substrategy",
                "currency",
                "quantity",
                "qty_premium",
            ],
        labels=trades.COLUMN_LABELS,
        filter_model={"qty_premium": {"type": "lessThan", "filter": 0}},
        selected_substrategies=["Safe strategy", "=unsafe formula"],
    )

    workbook = load_workbook(BytesIO(payload), data_only=False)
    assert workbook.sheetnames == ["Trades", "Snapshot Metadata"]
    sheet = workbook["Trades"]
    assert isinstance(sheet["A2"].value, datetime)
    assert sheet["C2"].value == "USD"
    assert isinstance(sheet["D2"].value, float)
    assert isinstance(sheet["E2"].value, float)
    assert sheet["B3"].value == "'=unsafe formula"
    assert sheet.auto_filter.ref == sheet.dimensions

    metadata = {
        row[0].value: row[1].value
        for row in workbook["Snapshot Metadata"].iter_rows(min_row=2)
    }
    assert metadata["Source"] == snapshot.metadata()["source"]
    assert metadata["Snapshot row count"] == 2
    assert metadata["Exported row count"] == 2
    assert "execution-price basis" in metadata["Premium convention"]
    assert metadata["Native currencies"] == "USD"
    assert "qty_premium" in metadata["Grid filters"]
    assert "eight decimal places" in metadata["Model Greek units"]
    assert "six decimal places" in metadata["Quantity Greek units"]
    assert "two decimal places" in metadata["Quantity monetary units"]
    assert "Risk currencies" not in metadata


def test_workbook_preserves_decimal_monetary_and_risk_values_as_numbers():
    payload = trade_ledger.build_trade_workbook(
        [
            {
                "delta_s1": Decimal("0.12345678"),
                "qty_pnl": Decimal("123.45"),
                "qty_delta_asset_a": Decimal("-987.654321"),
            }
        ],
        {},
        columns=["delta_s1", "qty_pnl", "qty_delta_asset_a"],
        labels=trades.COLUMN_LABELS,
    )

    workbook = load_workbook(BytesIO(payload), data_only=False)
    trades_sheet = workbook["Trades"]

    assert trades_sheet["A2"].data_type == "n"
    assert trades_sheet["A2"].value == 0.12345678
    assert trades_sheet["B2"].data_type == "n"
    assert trades_sheet["B2"].value == 123.45
    assert trades_sheet["B2"].number_format == "#,##0.00"
    assert trades_sheet["C2"].data_type == "n"
    assert trades_sheet["C2"].value == -987.654321
    assert trades_sheet["C2"].number_format == "#,##0.000000"
    assert trades_sheet["A2"].number_format == "#,##0.00000000"


def test_page_has_one_h1_one_wide_grid_and_no_summary_detail_selector():
    components = list(_component_tree(trades.layout))
    names = [type(component).__name__ for component in components]
    ids = {
        getattr(component, "id", None)
        for component in components
        if getattr(component, "id", None)
    }

    assert names.count("H1") == 1
    assert names.count("AgGrid") == 1
    assert "RadioItems" not in names
    assert "trades-view-selector" not in ids
    assert "trades-table" in ids
    heading = next(
        component
        for component in components
        if type(component).__name__ == "H1"
    )
    assert heading.children == "Trade ledger"
    assert "trades-visually-hidden-heading" in heading.className.split()
    assert not any(
        getattr(component, "children", None)
        == (
            "Every active option trade leg with its booked economics and "
            "published valuation and Greeks for the selected COB."
        )
        for component in components
    )


def test_trades_filters_reuse_the_greeks_sticky_toolbar_contract():
    toolbar = next(
        component
        for component in _component_tree(trades.layout)
        if "trades-monitor-controls"
        in getattr(component, "className", "").split()
    )
    selector_row = next(
        component
        for component in _component_tree(toolbar)
        if "trades-monitor-selector-row"
        in getattr(component, "className", "").split()
    )
    ids = {
        getattr(component, "id", None)
        for component in _component_tree(selector_row)
        if getattr(component, "id", None)
    }
    labels = [
        component.children
        for component in _component_tree(selector_row)
        if "inline-filter-label"
        in getattr(component, "className", "").split()
    ]

    assert "greeks-sticky-filter-bar" in toolbar.className.split()
    assert labels == ["COB Date", "Strategies"]
    assert ids == {
        "trades-date-dropdown",
        "trades-strategy-dropdown",
        "trades-source-status-mount",
        "trades-dashboard-source-status-inline",
    }
    strategy_dropdown = next(
        component
        for component in _component_tree(selector_row)
        if getattr(component, "id", None) == "trades-strategy-dropdown"
    )
    assert "greeks-compact-multi-dropdown" in strategy_dropdown.className.split()
    assert "greeks-inline-source-status" in selector_row.children[2].className.split()
    assert not any(
        getattr(component, "id", None) == "export-trades-table-btn"
        for component in _component_tree(toolbar)
    )

    section_actions = next(
        component
        for component in _component_tree(trades.layout)
        if "trades-section-actions"
        in getattr(component, "className", "").split()
    )
    assert any(
        getattr(component, "id", None) == "export-trades-table-btn"
        for component in _component_tree(section_actions)
    )


def test_grid_contains_required_groups_and_native_numeric_columns():
    groups = [definition["headerName"] for definition in trades.TRADES_COLUMN_DEFS]
    fields = dict(_column_fields(trades.TRADES_COLUMN_DEFS))

    assert groups == [
        "Trade",
        "Position and premium",
        "Valuation",
        "Position Greeks (original fields)",
        "Model Greeks per unit (original fields)",
        "Contract legs",
        "Market and model inputs",
        "Valuation lineage",
    ]
    for field in (
        "trade_date",
        "premium",
        "qty_premium",
        "qty_value",
        "qty_pnl",
        *trade_ledger.ORIGINAL_MODEL_GREEK_COLUMNS,
        *trade_ledger.ORIGINAL_POSITION_GREEK_COLUMNS,
    ):
        assert field in fields
    assert fields["qty_premium"]["cellDataType"] == "number"
    assert fields["qty_rho"]["filter"] == "agNumberColumnFilter"
    assert fields["trade_date"]["pinned"] == "left"
    assert not any(field.startswith("model_") for field in fields)
    assert "risk_currency" not in fields
    assert "smile_call_delta_used" not in fields
    assert "smile_call_delta_used" not in trades.EXPORT_COLUMNS
    assert "vol_swaption_discount" not in trade_ledger.OUTPUT_COLUMNS
    assert "vol_swaption_discount" not in fields
    assert "vol_swaption_discount" not in trades.EXPORT_COLUMNS
    for field in (
        "forward_source_name",
        "forward_source_cob_date",
        "volatility_source_name",
        "volatility_source_cob_date",
        "volatility_method",
    ):
        assert field not in trade_ledger.OUTPUT_COLUMNS
        assert field not in fields
        assert field not in trades.EXPORT_COLUMNS


def test_grid_numeric_formatters_apply_locale_thousands_separators():
    fields = dict(_column_fields(trades.TRADES_COLUMN_DEFS))

    for field in (
        "quantity",
        "premium",
        "qty_premium",
        "qty_value",
        "qty_delta_asset_a",
        "volatility_used",
        "valuation_revision",
    ):
        formatter = fields[field]["valueFormatter"]["function"]
        assert "d3.format(',." in formatter
        assert fields[field]["cellDataType"] == "number"

    assert "d3.format(',.2f')" in fields["quantity"]["valueFormatter"]["function"]
    assert "d3.format(',.4f')" in fields["premium"]["valueFormatter"]["function"]
    assert "d3.format(',.2f')" in fields["qty_pnl"]["valueFormatter"]["function"]
    assert "d3.format(',.6f')" in fields["qty_delta_asset_a"]["valueFormatter"]["function"]
    assert "d3.format(',.8f')" in fields["delta_s1"]["valueFormatter"]["function"]


def test_table_callback_surfaces_database_errors(monkeypatch):
    monkeypatch.setattr(trades, "get_database_engine", lambda required=False: object())

    def fail_load(*args, **kwargs):
        raise RuntimeError("database offline")

    monkeypatch.setattr(trades, "load_trade_snapshot", fail_load)
    result = trades.update_trades_table(
        "2026-07-27",
        ["JKM call"],
        0,
        {"error": ""},
        {"error": ""},
    )

    assert result[0] == []
    assert "unavailable" in result[2].lower()
    assert "failed" in result[4].lower()
    assert result[5] == trades.ERROR_STYLE_VISIBLE


def test_substrategy_discovery_error_is_propagated_to_visible_page_state(
    monkeypatch,
):
    monkeypatch.setattr(trades, "get_database_engine", lambda required=False: object())

    def fail_discovery(*args, **kwargs):
        raise RuntimeError("database offline")

    monkeypatch.setattr(trades, "get_substrategies", fail_discovery)
    options, selected, state = trades.update_strategy_options(
        "2026-07-27",
        0,
        [],
    )
    result = trades.update_trades_table(
        "2026-07-27",
        [],
        0,
        {"error": ""},
        state,
    )

    assert options == []
    assert selected == []
    assert state["error"]
    assert result[0] == []
    assert result[5] == trades.ERROR_STYLE_VISIBLE


def test_export_uses_virtual_grid_rows_in_their_current_order(monkeypatch):
    captured = {}

    def fake_build(records, metadata, **kwargs):
        captured["records"] = records
        captured["metadata"] = metadata
        captured["kwargs"] = kwargs
        return b"workbook"

    monkeypatch.setattr(trades, "build_trade_workbook", fake_build)
    virtual_rows = [
        {"_trade_key": "second", "substrategy": "B"},
        {"_trade_key": "first", "substrategy": "A"},
    ]
    download, status = trades.export_trades_table(
        1,
        "2026-07-27",
        ["A", "B"],
        virtual_rows,
        [{"_trade_key": "first", "substrategy": "A"}],
        {"substrategy": {"filter": "A"}},
        {"cob_date": "2026-07-27"},
    )

    assert captured["records"] == virtual_rows
    assert captured["kwargs"]["filter_model"] == {
        "substrategy": {"filter": "A"}
    }
    assert download["filename"] == "trade_ledger_2026-07-27_2_rows.xlsx"
    assert status == "Exported 2 rows."
